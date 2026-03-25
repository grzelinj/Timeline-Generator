from openpyxl import load_workbook
from openpyxl.styles import Font
from datetime import datetime, timedelta
import re

# -----------------------------------------
# 1. LISTA ŚWIĄT – UK + Barcelona (2026)
# -----------------------------------------

uk_holidays_2026 = {
    datetime(2026, 4, 3),
    datetime(2026, 4, 6),
    datetime(2026, 5, 4),
    datetime(2026, 5, 25),
    datetime(2026, 8, 31),
    datetime(2026, 12, 25),
    datetime(2026, 12, 28),
}

barcelona_holidays_2026 = {
    datetime(2026, 1, 1),
    datetime(2026, 1, 6),
    datetime(2026, 4, 3),
    datetime(2026, 4, 6),
    datetime(2026, 5, 1),
    datetime(2026, 5, 25),
    datetime(2026, 6, 24),
    datetime(2026, 8, 15),
    datetime(2026, 9, 11),
    datetime(2026, 9, 24),
    datetime(2026, 10, 12),
    datetime(2026, 12, 8),
    datetime(2026, 12, 25),
    datetime(2026, 12, 26),
}

ALL_HOLIDAYS = uk_holidays_2026.union(barcelona_holidays_2026)

# -----------------------------------------
# 2. FUNKCJA DNI ROBOCZYCH
# -----------------------------------------

def add_business_days(start_date, days):
    step = 1 if days > 0 else -1
    remaining = abs(days)
    d = start_date

    while remaining > 0:
        d += timedelta(days=step)
        if d.weekday() < 5 and d not in ALL_HOLIDAYS:
            remaining -= 1
    return d

# -----------------------------------------
# 3. PARSOWANIE OFFSETU
# -----------------------------------------

def parse_offset(text):
    if not isinstance(text, str):
        return None

    match = re.search(r"Vesting Date\s*([+\-\u2212\u2013\u2014])\s*(\d+)\s*days", text)

    if not match:
        return None

    sign = match.group(1)
    value = int(match.group(2))

    if sign in ["-", "−", "–", "—"]:
        return -value
    if sign == "+":
        return value

    return None

# -----------------------------------------
# 4. GENERATOR DUE DATES
# -----------------------------------------

def generate_due_dates(vesting_date_str, template_path):
    vesting = datetime.strptime(vesting_date_str, "%d/%m/%Y")
    wb = load_workbook(template_path)
    ws = wb["Tasklist"]

    for row in range(1, ws.max_row + 1):
        text = ws[f"A{row}"].value
        days = parse_offset(text)

        if days is not None:
            new_due = add_business_days(vesting, days)
            cell = ws[f"B{row}"]
            cell.value = new_due.strftime("%A, %d/%m/%Y")

    # ✅ POGUBIENIE WIERSZY 25 i 28
    for bold_row in [25, 28]:
        for col in range(1, ws.max_column + 1):
            ws[f"{chr(64+col)}{bold_row}"].font = Font(bold=True)

    safe_date = vesting_date_str.replace("/", ".")
    new_name = f"Tate & Lyle Vesting Task List Vest {safe_date}.xlsx"
    wb.save(new_name)
    return new_name

# -----------------------------------------
# 5. URUCHOMIENIE
# -----------------------------------------

if __name__ == "__main__":
    generate_due_dates("13/05/2026", "template.xlsx")


    
# -----------------------------------------
# MOJE NOTATKI (Jan)
#Otwierasz miniconde i aktywujesz srodowisko:  conda activate od_zera_do_ai

# Wskazanie folderu:
#cd "C:\Users\grzelinskij\Desktop\JG\od_zera_do_ai\JG apps\Project 1 - Timeline Generator"

#Aktywowanie Pythona w Terminalu:
#"%USERPROFILE%\AppData\Local\miniconda3\envs\od_zera_do_ai\python.exe" --version

#POTEM TO:
#cd "C:\Users\grzelinskij\Desktop\JG\od_zera_do_ai\JG apps\Project 1 - Timeline Generator"

# jak zrobisz wszytskie updaty, to pozniej w 
#git add .
#git commit -m "update timeline logic"
#git push


# - dodać obsługę wielu template
# - zintegrować z API EquatePlus?
# - sprawdzić ścieżkę Pythona
# - dodać logikę dla PSP / RSA / EXCO
# -----------------------------------------
